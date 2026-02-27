from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.core.paginator import Paginator, EmptyPage, PageNotAnInteger
from django.http import JsonResponse, HttpResponseForbidden, HttpResponse
from django.core.exceptions import ValidationError
from django.db.models import Q, Sum
from .models import Activity, ActivityAttachment
from .forms import ActivityForm, BulkActionForm, ActivityAttachmentForm
from accounts.models import Cluster, User
from masters.models import Funder, ActivityStatus, Currency, ProcurementType
from audit.models import AuditLog
import json
from datetime import datetime, date
from decimal import Decimal


MONTH_SHORT_NAMES = {
    1: 'Jan',
    2: 'Feb',
    3: 'Mar',
    4: 'Apr',
    5: 'May',
    6: 'Jun',
    7: 'Jul',
    8: 'Aug',
    9: 'Sep',
    10: 'Oct',
    11: 'Nov',
    12: 'Dec',
}


# Permission helper functions
def has_role(user, role_name):
    """Check if user has a specific role or is superuser"""
    return user.is_superuser or user.groups.filter(name=role_name).exists()


def can_view_activities(user):
    """Check if user can view activities"""
    return any(has_role(user, role) for role in ['System Admin', 'Data Manager', 'Activity Manager', 'Project Coordinator', 'Viewer'])


def can_edit_activities(user):
    """Check if user can edit activities"""
    return any(has_role(user, role) for role in ['System Admin', 'Data Manager', 'Activity Manager', 'Project Coordinator'])


def can_manage_activities(user):
    """Check if user can create, delete, and bulk manage activities"""
    return any(has_role(user, role) for role in ['System Admin', 'Data Manager', 'Activity Manager'])


def can_manage_users(user):
    """Check if user can manage users"""
    return any(has_role(user, role) for role in ['System Admin', 'User Manager'])


def get_responsible_officer_options():
    """Return active officers and their cluster memberships for dynamic filtering in forms."""
    officers = User.objects.filter(is_active=True).prefetch_related('clusters').order_by('first_name', 'last_name', 'username')
    options = []
    for officer in officers:
        full_name = officer.get_full_name().strip() if officer.get_full_name() else officer.username
        options.append({
            'id': officer.id,
            'name': full_name,
            'cluster_ids': list(officer.clusters.values_list('id', flat=True)),
        })
    return options


def get_procurement_type_options():
    """Return active procurement types for create/edit JS UI."""
    return list(
        ProcurementType.objects.filter(active=True)
        .order_by('name')
        .values('code', 'name')
    )


def safe_parse_json_list(raw_value):
    """Safely parse JSON list payloads and return [] on invalid input."""
    if not raw_value:
        return []
    try:
        parsed = json.loads(raw_value)
        return parsed if isinstance(parsed, list) else []
    except (TypeError, ValueError):
        return []


def can_use_cluster_filter(user):
    """Check if user should see the cluster filter"""
    if not user or not user.is_authenticated:
        return False
    if user.is_superuser or user.is_staff:
        return True
    return any(has_role(user, role) for role in ['System Admin', 'Data Manager', 'Project Coordinator', 'Viewer'])


def get_user_role_category(user):
    """Categorize user into role categories for access control"""
    if user.is_superuser or has_role(user, 'System Admin'):
        return 'admin'
    elif has_role(user, 'Data Manager') or has_role(user, 'Viewer'):
        return 'viewer'
    elif has_role(user, 'Activity Manager'):
        return 'manager'
    elif has_role(user, 'Project Coordinator'):
        return 'coordinator'
    return 'none'


def apply_role_based_filters(queryset, user):
    """Apply role-based filtering to queryset based on user role"""
    role_category = get_user_role_category(user)

    if role_category == 'manager':
        # Activity Manager: Only see activities from their clusters
        user_clusters = user.clusters.all()
        if user_clusters.exists():
            queryset = queryset.filter(clusters__in=user_clusters).distinct()
        else:
            # No clusters assigned, show no data
            queryset = queryset.none()
    # For 'admin', 'viewer', and 'coordinator': Show all data (no filtering)
    
    return queryset



@login_required
def activities_list(request):
    if not can_view_activities(request.user):
        return HttpResponseForbidden("You do not have permission to view activities")
    
    qs = (
        Activity.objects.filter(deleted=False).select_related('status', 'currency', 'responsible_officer')
        .prefetch_related('clusters', 'funders')
        .order_by('-year', 'activity_id')
    )

    # Get unique years from activities for dropdown, sorted descending
    available_years = sorted(Activity.objects.filter(deleted=False).values_list('year', flat=True).distinct(), reverse=True)
    
    # Filters
    year = request.GET.get('year')
    cluster = request.GET.get('cluster')
    funder = request.GET.get('funder')
    status = request.GET.get('status')
    quarter = request.GET.get('quarter')
    month = request.GET.get('month')
    search_query = request.GET.get('q')  # Search query
    procurement_status = request.GET.get('procurement_status', 'all')  # New filter
    recurring_filter = request.GET.get('recurring', 'all')  # New filter
    show_all_years = request.GET.get('all_years') == '1'  # New parameter

    # Set default year to current year if no year filter and not showing all years
    current_year = date.today().year
    if not year and not show_all_years and current_year in available_years:
        year = str(current_year)
        qs = qs.filter(year=current_year)
    elif year:
        try:
            qs = qs.filter(year=int(year))
        except ValueError:
            pass
    
    # Apply role-based filtering
    role_category = get_user_role_category(request.user)
    qs = apply_role_based_filters(qs, request.user)
    qs = qs.filter(deleted=False)
    
    # Determine which filters to show based on user role
    show_cluster_filter = can_use_cluster_filter(request.user)
    show_funder_filter = True
    
    needs_distinct = False

    if cluster:
        if cluster.isdigit():
            qs = qs.filter(clusters__id=int(cluster))
        else:
            qs = qs.filter(clusters__short_name=cluster)
        needs_distinct = True

    if funder:
        if funder.isdigit():
            qs = qs.filter(funders__id=int(funder))
        else:
            qs = qs.filter(funders__code=funder)
        needs_distinct = True

    if status:
        if status.isdigit():
            qs = qs.filter(status__id=int(status))
        else:
            qs = qs.filter(status__name=status)

    if quarter:
        try:
            quarter_num = int(quarter)
            if quarter_num in [1, 2, 3, 4]:
                qs = qs.filter(quarter=quarter_num)
        except ValueError:
            pass

    if month:
        try:
            qs = qs.filter(planned_month__month=int(month))
        except ValueError:
            pass

    if search_query:
        qs = qs.filter(Q(name__icontains=search_query) | Q(activity_id__icontains=search_query))

    if needs_distinct:
        qs = qs.distinct()
    
    # Procurement filter
    if procurement_status == 'procurement_only':
        qs = qs.filter(is_procurement=True)
    elif procurement_status == 'has_procurement':
        qs = qs.filter(Q(is_procurement=True) | Q(has_partial_procurement=True))
    elif procurement_status == 'non_procurement':
        qs = qs.filter(is_procurement=False, has_partial_procurement=False)
    
    # Recurring filter
    if recurring_filter == 'recurring_only':
        qs = qs.filter(is_recurring=True)
    elif recurring_filter == 'generated':
        qs = qs.filter(generated_from_recurrence=True)
    elif recurring_filter == 'non_recurring':
        qs = qs.filter(is_recurring=False, generated_from_recurrence=False)

    # Calculate summary statistics for filtered activities
    total_activities = qs.count()
    total_budget = qs.aggregate(Sum('total_budget'))['total_budget__sum'] or 0
    total_disbursed = qs.aggregate(Sum('disbursed_amount'))['disbursed_amount__sum'] or 0
    total_procurement_value = qs.aggregate(Sum('procurement_amount'))['procurement_amount__sum'] or 0
    procurement_activities_count = qs.filter(Q(is_procurement=True) | Q(has_partial_procurement=True)).count()

    # Sorting
    sort = request.GET.get('sort')
    direction = request.GET.get('dir', 'asc')
    allowed_sorts = {
        'activity_id': 'activity_id',
        'name': 'name',
        'year': 'year',
        'cluster': 'clusters__short_name',
        'status': 'status__name',
        'budget': 'total_budget',
    }
    if sort in allowed_sorts:
        sort_field = allowed_sorts[sort]
        if direction == 'desc':
            qs = qs.order_by('-' + sort_field)
        else:
            qs = qs.order_by(sort_field)

    # Build URLs for year toggle
    current_params = request.GET.copy()
    if show_all_years:
        current_params.pop('all_years', None)
        show_current_year_url = request.path + ('?' + current_params.urlencode() if current_params else '')
        show_all_years_url = request.get_full_path()
    else:
        current_params['all_years'] = '1'
        show_all_years_url = request.path + '?' + current_params.urlencode()
        show_current_year_url = request.get_full_path()

    # Determine available months based on selected quarter
    if quarter:
        try:
            quarter_num = int(quarter)
            if quarter_num == 1:
                available_months = [1, 2, 3]
            elif quarter_num == 2:
                available_months = [4, 5, 6]
            elif quarter_num == 3:
                available_months = [7, 8, 9]
            elif quarter_num == 4:
                available_months = [10, 11, 12]
            else:
                available_months = list(range(1, 13))
        except ValueError:
            available_months = list(range(1, 13))
    else:
        available_months = list(range(1, 13))

    available_month_options = [
        {'value': month_num, 'label': MONTH_SHORT_NAMES.get(month_num, str(month_num))}
        for month_num in available_months
    ]

    # Get filtered data for visibility logic
    filtered_clusters = qs.values('clusters').distinct().count()
    filtered_funders = qs.values('funders').distinct().count()

    context = {
        'activities': qs,
        'clusters': Cluster.objects.all(),
        'funders': Funder.objects.all(),
        'statuses': ActivityStatus.objects.all(),
        'users': User.objects.filter(is_active=True),
        'years': available_years,
        'available_months': available_months,
        'available_month_options': available_month_options,
        'filters': {
            'year': year,
            'cluster': cluster,
            'funder': funder,
            'status': status,
            'quarter': quarter,
            'month': month,
            'q': search_query,
            'procurement_status': procurement_status,
            'recurring': recurring_filter,
            'all_years': show_all_years,
        },
        'sort': sort,
        'dir': direction,
        'can_create': can_manage_activities(request.user),
        'can_edit': can_edit_activities(request.user),
        'show_all_years_url': request.get_full_path() + ('&' if request.GET else '?') + 'all_years=1',
        'show_current_year_url': request.get_full_path().replace('&all_years=1', '').replace('?all_years=1', '').replace('all_years=1&', ''),
        'user_role': role_category,
        'show_cluster_filter': show_cluster_filter,
        'show_funder_filter': show_funder_filter,
        'total_activities': total_activities,
        'total_budget': total_budget,
        'total_disbursed': total_disbursed,
        'total_procurement_value': total_procurement_value,
        'procurement_activities_count': procurement_activities_count,
    }
    return render(request, 'activities/list.html', context)


@login_required
def activity_detail(request, pk):
    a = get_object_or_404(Activity, pk=pk, deleted=False)
    audit_logs = AuditLog.objects.filter(
        activity_id=pk
    ).select_related('user').order_by('-timestamp')[:50]
    
    can_edit = can_edit_activities(request.user)
    can_delete = can_manage_activities(request.user)
    available_years = sorted(Activity.objects.filter(deleted=False).values_list('year', flat=True).distinct(), reverse=True)
    
    context = {
        'activity': a,
        'audit_logs': audit_logs,
        'all_users': User.objects.filter(is_active=True, clusters__in=a.clusters.all()).distinct().order_by('first_name', 'last_name', 'username'),
        'all_statuses': ActivityStatus.objects.all().order_by('name'),
        'users': User.objects.filter(is_active=True),
        'statuses': ActivityStatus.objects.all(),
        'clusters': Cluster.objects.all(),
        'funders': Funder.objects.all(),
        'currencies': Currency.objects.all(),
        'years': available_years,
        'can_edit': can_edit,
        'can_delete': can_delete,
    }
    return render(request, 'activities/detail.html', context)


@login_required
def edit_activity(request, pk):
    activity = get_object_or_404(Activity, pk=pk, deleted=False)
    
    if not can_edit_activities(request.user):
        return HttpResponseForbidden("You do not have permission to edit activities")
    
    if request.method == 'GET':
        # Render edit form
        form = ActivityForm(instance=activity)
        procurement_types = ProcurementType.objects.filter(active=True).order_by('name')
        context = {
            'form': form,
            'activity': activity,
            'title': 'Edit Activity',
            'procurement_types': procurement_types,
            'officer_options': get_responsible_officer_options(),
            'procurement_type_options': get_procurement_type_options(),
            'existing_procurement_breakdowns': activity.procurement_breakdowns or [],
        }
        return render(request, 'activities/create.html', context)
    
    elif request.method == 'POST':
        # Check if this is a form submission or JSON (inline edit)
        content_type = request.content_type
        
        if 'application/json' in content_type:
            # Handle inline JSON updates
            try:
                data = json.loads(request.body)
                field = data.get('field')
                value = data.get('value')
                
                # Track old value for audit log
                if field in ['clusters', 'funders']:
                    old_value = list(getattr(activity, field).all())
                else:
                    old_value = getattr(activity, field, None)
                
                # Update field based on type
                if field == 'responsible_officer':
                    if value:
                        activity.responsible_officer = User.objects.get(id=int(value))
                    else:
                        activity.responsible_officer = None
                elif field == 'status':
                    activity.status = ActivityStatus.objects.get(id=int(value))
                elif field == 'notes':
                    activity.notes = value or ''
                elif field == 'name':
                    activity.name = value
                elif field == 'year':
                    activity.year = int(value)
                elif field == 'clusters':
                    activity.clusters.set([Cluster.objects.get(id=int(id)) for id in value])
                elif field == 'funders':
                    activity.funders.set([Funder.objects.get(id=int(id)) for id in value])
                elif field == 'total_budget':
                    activity.total_budget = Decimal(str(value)) if value not in [None, ''] else Decimal('0')
                elif field == 'disbursed_amount':
                    activity.disbursed_amount = Decimal(str(value)) if value not in [None, ''] else Decimal('0')
                elif field == 'currency':
                    activity.currency = Currency.objects.get(id=int(value))
                elif field == 'planned_month':
                    activity.planned_month = datetime.fromisoformat(value).date() if value else None
                # Procurement fields
                elif field == 'is_procurement':
                    activity.is_procurement = value.lower() == 'true' if isinstance(value, str) else bool(value)
                elif field == 'procurement_type':
                    activity.procurement_type = value if value else None
                elif field == 'procurement_amount':
                    activity.procurement_amount = float(value) if value else None
                # Recurrence fields (only allow editing if this is a template, not a generated instance)
                elif field == 'is_recurring':
                    if activity.generated_from_recurrence:
                        return JsonResponse({'error': 'Cannot mark a generated instance as recurring'}, status=400)
                    activity.is_recurring = value.lower() == 'true' if isinstance(value, str) else bool(value)
                elif field == 'recurrence_pattern':
                    if activity.generated_from_recurrence:
                        return JsonResponse({'error': 'Cannot edit recurrence of a generated instance'}, status=400)
                    activity.recurrence_pattern = value if value else None
                elif field == 'recurrence_interval':
                    if activity.generated_from_recurrence:
                        return JsonResponse({'error': 'Cannot edit recurrence of a generated instance'}, status=400)
                    activity.recurrence_interval = int(value) if value else 1
                elif field == 'recurrence_end_date':
                    if activity.generated_from_recurrence:
                        return JsonResponse({'error': 'Cannot edit recurrence of a generated instance'}, status=400)
                    activity.recurrence_end_date = datetime.fromisoformat(value).date() if value else None
                else:
                    return JsonResponse({'error': 'Invalid field'}, status=400)
                
                # Validate and save
                activity.clean()
                activity.save()
                
                # Create audit log
                if field in ['clusters', 'funders']:
                    new_value = list(getattr(activity, field).all())
                    old_display = ', '.join([str(obj) for obj in old_value]) if old_value else 'None'
                    new_display = ', '.join([str(obj) for obj in new_value]) if new_value else 'None'
                elif field == 'planned_month':
                    old_display = old_value.strftime('%Y-%m-%d') if old_value else 'N/A'
                    new_display = getattr(activity, field).strftime('%Y-%m-%d') if getattr(activity, field) else 'N/A'
                else:
                    new_value = getattr(activity, field, None)
                    old_display = str(old_value)[:100] if old_value is not None else 'N/A'
                    new_display = str(new_value)[:100] if new_value is not None else 'N/A'
                
                AuditLog.objects.create(
                    user=request.user,
                    action=f'Activity {field} changed',
                    object_repr=str(activity),
                    activity_id=activity.id,
                    change_description=f'{field}: {old_display} → {new_display}'
                )
                
                return JsonResponse({'success': True, 'new_value': new_display})
            except ValidationError as e:
                return JsonResponse({'error': str(e)}, status=400)
            except Exception as e:
                return JsonResponse({'error': str(e)}, status=500)
        else:
            # Handle form submission
            form = ActivityForm(request.POST, instance=activity)
            if form.is_valid():
                updated_activity = form.save(commit=False)
                
                # Handle procurement breakdowns JSON
                procurement_data = request.POST.get('procurement_breakdowns', '')
                if procurement_data:
                    try:
                        updated_activity.procurement_breakdowns = json.loads(procurement_data)
                    except json.JSONDecodeError:
                        updated_activity.procurement_breakdowns = None
                
                updated_activity.save()
                form.save_m2m()  # Save many-to-many relationships
                
                AuditLog.objects.create(
                    user=request.user,
                    action='Activity updated',
                    object_repr=str(updated_activity),
                    activity_id=updated_activity.id,
                    change_description='Activity edited via form'
                )
                
                messages.success(request, f'Activity "{updated_activity.name}" updated successfully!')
                return redirect('activity_detail', pk=updated_activity.pk)
            else:
                context = {
                    'form': form,
                    'activity': activity,
                    'title': 'Edit Activity',
                    'procurement_types': ProcurementType.objects.filter(active=True).order_by('name'),
                    'officer_options': get_responsible_officer_options(),
                    'procurement_type_options': get_procurement_type_options(),
                    'existing_procurement_breakdowns': (
                        safe_parse_json_list(request.POST.get('procurement_breakdowns'))
                        or (form.instance.procurement_breakdowns or [])
                    ),
                }
                return render(request, 'activities/create.html', context)
    
    return HttpResponseForbidden("Method not allowed")


@login_required
def create_activity(request):
    """Create a new activity (Data Manager, System Admin only)"""
    if not can_manage_activities(request.user):
        return HttpResponseForbidden("You do not have permission to create activities")
    
    if request.method == 'POST':
        form = ActivityForm(request.POST)
        if form.is_valid():
            activity = form.save(commit=False)
            
            # Handle procurement breakdowns JSON
            procurement_data = request.POST.get('procurement_breakdowns', '')
            if procurement_data:
                try:
                    activity.procurement_breakdowns = json.loads(procurement_data)
                except json.JSONDecodeError:
                    activity.procurement_breakdowns = None
            
            activity.save()
            form.save_m2m()  # Save many-to-many relationships
            
            AuditLog.objects.create(
                user=request.user,
                action='Activity created',
                object_repr=str(activity),
                activity_id=activity.id,
                change_description='New activity created'
            )
            
            messages.success(request, f'Activity "{activity.name}" created successfully!')
            return redirect('activity_detail', pk=activity.pk)
    else:
        form = ActivityForm()
    
    procurement_types = ProcurementType.objects.filter(active=True).order_by('name')
    context = {
        'form': form,
        'title': 'Create New Activity',
        'procurement_types': procurement_types,
        'officer_options': get_responsible_officer_options(),
        'procurement_type_options': get_procurement_type_options(),
        'existing_procurement_breakdowns': (
            safe_parse_json_list(request.POST.get('procurement_breakdowns'))
            if request.method == 'POST' else []
        ),
    }
    return render(request, 'activities/create.html', context)


@login_required
def delete_activity(request, pk):
    """Soft delete an activity - mark as deleted (Data Manager, System Admin only)"""
    activity = get_object_or_404(Activity, pk=pk, deleted=False)
    
    if not can_manage_activities(request.user):
        messages.error(request, 'Permission denied')
        return redirect('activities_list')
    
    if request.method == 'POST':
        activity.deleted = True
        activity.save()
        
        AuditLog.objects.create(
            user=request.user,
            action='Activity deleted',
            object_repr=str(activity),
            activity_id=activity.id,
            change_description='Activity marked as deleted'
        )
        
        messages.success(request, f'Activity "{activity.name}" has been deleted successfully.')
        return redirect('activities_list')
    
    messages.error(request, 'Invalid request method')
    return redirect('activities_list')


@login_required
def bulk_action(request):
    """Handle bulk operations on activities (Data Manager, System Admin only)"""
    if not can_manage_activities(request.user):
        return JsonResponse({'error': 'Permission denied'}, status=403)
    
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            action = data.get('action')
            activity_ids = data.get('activity_ids', [])
            
            if not activity_ids:
                return JsonResponse({'error': 'No activities selected'}, status=400)
            
            qs = Activity.objects.filter(id__in=activity_ids, deleted=False)
            results = {'updated': 0, 'errors': []}
            
            if action == 'update_status':
                status_id = data.get('status_id')
                if not status_id:
                    return JsonResponse({'error': 'Status not provided'}, status=400)
                
                try:
                    status = ActivityStatus.objects.get(id=status_id)
                    updated = qs.update(status=status)
                    results['updated'] = updated
                    
                    for activity in qs:
                        AuditLog.objects.create(
                            user=request.user,
                            action='Bulk status update',
                            object_repr=str(activity),
                            activity_id=activity.id,
                            change_description=f'Status changed to {status.name} via bulk action'
                        )
                except ActivityStatus.DoesNotExist:
                    return JsonResponse({'error': 'Status not found'}, status=400)
            
            elif action == 'delete':
                updated = qs.update(deleted=True)
                results['updated'] = updated
                
                for activity in qs:
                    AuditLog.objects.create(
                        user=request.user,
                        action='Bulk delete',
                        object_repr=str(activity),
                        activity_id=activity.id,
                        change_description='Activity marked as deleted via bulk action'
                    )
            
            elif action == 'assign_officer':
                officer_id = data.get('officer_id')
                if not officer_id:
                    return JsonResponse({'error': 'Officer not provided'}, status=400)
                
                try:
                    officer = User.objects.get(id=officer_id)
                    updated = qs.update(responsible_officer=officer)
                    results['updated'] = updated
                    
                    for activity in qs:
                        AuditLog.objects.create(
                            user=request.user,
                            action='Bulk assign officer',
                            object_repr=str(activity),
                            activity_id=activity.id,
                            change_description=f'Officer assigned to {officer.get_full_name()} via bulk action'
                        )
                except User.DoesNotExist:
                    return JsonResponse({'error': 'Officer not found'}, status=400)
            
            else:
                return JsonResponse({'error': 'Invalid action'}, status=400)
            
            return JsonResponse({'success': True, 'results': results})
        
        except json.JSONDecodeError:
            return JsonResponse({'error': 'Invalid JSON'}, status=400)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


# ============================================================================
# ATTACHMENT MANAGEMENT VIEWS
# ============================================================================

@login_required
def upload_attachment(request, pk):
    """Upload a new attachment to an activity"""
    if not can_edit_activities(request.user):
        return HttpResponseForbidden("You do not have permission to upload attachments")
    
    activity = get_object_or_404(Activity, pk=pk, deleted=False)
    
    if request.method == 'POST':
        form = ActivityAttachmentForm(request.POST, request.FILES)
        if form.is_valid():
            try:
                attachment = ActivityAttachment.upload_new_version(
                    activity=activity,
                    file=request.FILES['file'],
                    document_type=form.cleaned_data['document_type'],
                    description=form.cleaned_data.get('description', ''),
                    uploaded_by=request.user
                )
                
                # Log audit trail
                AuditLog.objects.create(
                    user=request.user,
                    action='Attachment uploaded',
                    object_repr=str(activity),
                    activity_id=activity.id,
                    change_description=f'Document "{attachment.filename}" (v{attachment.version}) uploaded'
                )
                
                return JsonResponse({
                    'success': True,
                    'message': f'Document uploaded successfully (v{attachment.version})',
                    'attachment_id': attachment.id
                })
            except Exception as e:
                return JsonResponse({'success': False, 'error': str(e)}, status=400)
        else:
            errors = {field: str(errors[0]) for field, errors in form.errors.items()}
            return JsonResponse({'success': False, 'errors': errors}, status=400)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@login_required
def delete_attachment(request, pk):
    """Delete (soft delete) an attachment"""
    if not can_edit_activities(request.user):
        return HttpResponseForbidden("You do not have permission to delete attachments")
    
    if request.method == 'POST':
        attachment = get_object_or_404(ActivityAttachment, pk=pk)
        activity = attachment.activity
        filename = attachment.filename
        
        try:
            # Soft delete the attachment
            attachment.is_deleted = True
            attachment.save()
            
            # Log audit trail
            AuditLog.objects.create(
                user=request.user,
                action='Attachment deleted',
                object_repr=str(activity),
                activity_id=activity.id,
                change_description=f'Document "{filename}" (v{attachment.version}) deleted'
            )
            
            return JsonResponse({
                'success': True,
                'message': 'Document deleted successfully'
            })
        except Exception as e:
            return JsonResponse({'success': False, 'error': str(e)}, status=400)
    
    return JsonResponse({'error': 'Method not allowed'}, status=405)


@login_required
def list_attachments(request, pk):
    """List all attachments for an activity with version history"""
    if not can_view_activities(request.user):
        return HttpResponseForbidden("You do not have permission to view attachments")
    
    activity = get_object_or_404(Activity, pk=pk, deleted=False)
    
    # Get all non-deleted attachments grouped by document type
    attachments = ActivityAttachment.objects.filter(
        activity=activity,
        is_deleted=False
    ).order_by('document_type', '-version')
    
    # Group by document type
    grouped = {}
    for attachment in attachments:
        doc_type = attachment.get_document_type_display()
        if doc_type not in grouped:
            grouped[doc_type] = []
        grouped[doc_type].append({
            'id': attachment.id,
            'filename': attachment.filename,
            'version': attachment.version,
            'is_latest': attachment.is_latest,
            'file_type': attachment.get_file_type_display(),
            'file_size': f"{attachment.file_size / 1024:.2f} KB",
            'uploaded_by': attachment.uploaded_by.get_full_name() if attachment.uploaded_by else 'Unknown',
            'uploaded_at': attachment.uploaded_at.strftime('%Y-%m-%d %H:%M'),
            'description': attachment.description,
            'download_url': attachment.file.url if attachment.file else '#'
        })
    
    return JsonResponse({
        'success': True,
        'attachments': grouped,
        'total_count': attachments.count()
    })


@login_required
def download_attachment(request, pk):
    """Download an attachment"""
    if not can_view_activities(request.user):
        return HttpResponseForbidden("You do not have permission to download attachments")
    
    attachment = get_object_or_404(ActivityAttachment, pk=pk)
    
    if attachment.is_deleted:
        return HttpResponseForbidden("This document has been deleted")
    
    if not attachment.file:
        return HttpResponseForbidden("File not found")
    
    try:
        # Log download in audit trail
        AuditLog.objects.create(
            user=request.user,
            action='Attachment downloaded',
            object_repr=str(attachment.activity),
            activity_id=attachment.activity.id,
            change_description=f'Document "{attachment.filename}" (v{attachment.version}) downloaded'
        )
        
        # Serve the file
        response = HttpResponse(attachment.file.read(), content_type='application/octet-stream')
        response['Content-Disposition'] = f'attachment; filename="{attachment.filename}"'
        return response
    except Exception as e:
        return HttpResponseForbidden(f"Error downloading file: {str(e)}")


@login_required
def get_attachment_versions(request, pk):
    """Get all versions of a specific document type"""
    if not can_view_activities(request.user):
        return HttpResponseForbidden("You do not have permission to view attachments")
    
    activity = get_object_or_404(Activity, pk=pk, deleted=False)
    document_type = request.GET.get('type', '')
    
    if not document_type:
        return JsonResponse({'error': 'Document type required'}, status=400)
    
    # Get all versions (including deleted)
    versions = ActivityAttachment.objects.filter(
        activity=activity,
        document_type=document_type
    ).order_by('-version')
    
    versions_data = []
    for v in versions:
        versions_data.append({
            'id': v.id,
            'version': v.version,
            'filename': v.filename,
            'file_size': f"{v.file_size / 1024:.2f} KB",
            'uploaded_by': v.uploaded_by.get_full_name() if v.uploaded_by else 'Unknown',
            'uploaded_at': v.uploaded_at.strftime('%Y-%m-%d %H:%M'),
            'is_latest': v.is_latest,
            'is_deleted': v.is_deleted,
            'description': v.description
        })
    
    return JsonResponse({
        'success': True,
        'document_type': document_type,
        'versions': versions_data
    })


@login_required
def update_activity_field(request, pk):
    """AJAX endpoint to update individual activity fields"""
    if request.method != 'POST':
        return JsonResponse({'success': False, 'error': 'Invalid request method'})
    
    activity = get_object_or_404(Activity, pk=pk, deleted=False)
    
    # Check permissions: must be data manager or responsible officer
    is_data_manager = has_role(request.user, 'Data Manager') or has_role(request.user, 'System Admin')
    is_responsible = activity.responsible_officer == request.user
    
    if not (is_data_manager or is_responsible):
        return JsonResponse({'success': False, 'error': 'You do not have permission to edit this activity'})
    
    field = request.POST.get('field')
    value = request.POST.get('value')
    
    try:
        old_value = None
        display_value = None
        
        if field == 'status':
            old_value = activity.status.name if activity.status else None
            status = get_object_or_404(ActivityStatus, pk=value)
            activity.status = status
            display_value = status.name
            
        elif field == 'responsible_officer':
            old_value = activity.responsible_officer.get_full_name() if activity.responsible_officer else 'Not assigned'
            if value:
                officer = get_object_or_404(User, pk=value)
                activity.responsible_officer = officer
                display_value = officer.get_full_name() or officer.username
            else:
                activity.responsible_officer = None
                display_value = 'Not assigned'
                
        elif field == 'notes':
            old_value = activity.notes
            activity.notes = value
            display_value = value or 'No notes provided'
            
        else:
            return JsonResponse({'success': False, 'error': 'Invalid field'})
        
        activity.save()
        
        # Log the change
        AuditLog.objects.create(
            user=request.user,
            activity_id=activity.id,
            action=f'Updated {field}',
            object_repr=str(activity),
            change_description=f'Changed from "{old_value}" to "{display_value}"'
        )
        
        return JsonResponse({
            'success': True,
            'field': field,
            'display_value': display_value
        })
        
    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


@login_required
def procurement_list(request):
    """View to list all activities with procurement"""
    if not can_view_activities(request.user):
        return HttpResponseForbidden("You do not have permission to view activities")
    
    qs = (
        Activity.objects.filter(Q(is_procurement=True) | Q(has_partial_procurement=True), deleted=False)
        .select_related('status', 'currency', 'responsible_officer')
        .prefetch_related('clusters', 'funders')
        .order_by('-year', 'activity_id')
    )

    # Apply role-based filtering
    role_category = get_user_role_category(request.user)
    qs = apply_role_based_filters(qs, request.user)
    
    # Determine which filters to show based on user role
    show_cluster_filter = can_use_cluster_filter(request.user)
    show_funder_filter = True

    status = request.GET.get('status')
    funder = request.GET.get('funder')
    cluster = request.GET.get('cluster')
    quarter = request.GET.get('quarter')
    procurement_type = request.GET.get('procurement_type', 'all')  # all | full | partial

    needs_distinct = False

    if status:
        if status.isdigit():
            qs = qs.filter(status__id=int(status))
        else:
            qs = qs.filter(status__name=status)

    if funder:
        if funder.isdigit():
            qs = qs.filter(funders__id=int(funder))
        else:
            qs = qs.filter(funders__code=funder)
        needs_distinct = True

    if cluster:
        if cluster.isdigit():
            qs = qs.filter(clusters__id=int(cluster))
        else:
            qs = qs.filter(clusters__short_name=cluster)
        needs_distinct = True

    if quarter:
        try:
            qs = qs.filter(quarter=int(quarter))
        except ValueError:
            pass

    if procurement_type == 'full':
        qs = qs.filter(is_procurement=True)
    elif procurement_type == 'partial':
        qs = qs.filter(has_partial_procurement=True)

    if needs_distinct:
        qs = qs.distinct()

    context = {
        'activities': qs,
        'clusters': Cluster.objects.all(),
        'funders': Funder.objects.all(),
        'statuses': ActivityStatus.objects.all(),
        'filters': {
            'status': status,
            'funder': funder,
            'cluster': cluster,
            'quarter': quarter,
            'procurement_type': procurement_type,
        },
        'user_role': role_category,
        'show_cluster_filter': show_cluster_filter,
        'show_funder_filter': show_funder_filter,
    }
    return render(request, 'activities/procurement_list.html', context)


@login_required
def procurement_detail(request, pk):
    """Detailed view of a single procurement activity"""
    if not can_view_activities(request.user):
        return HttpResponseForbidden("You do not have permission to view activities")
    
    activity = get_object_or_404(
        Activity.objects.select_related('status', 'currency', 'responsible_officer').prefetch_related('clusters', 'funders'),
        pk=pk, deleted=False
    )
    
    # Calculate totals from breakdown
    breakdown_total = 0
    if activity.procurement_breakdowns:
        for item in activity.procurement_breakdowns:
            breakdown_total += float(item.get('amount', 0) or 0)
    
    # Check if breakdown matches recorded amount (with small tolerance for floating point)
    procurement_amount = float(activity.procurement_amount or 0)
    amounts_match = abs(breakdown_total - procurement_amount) < 0.01
    
    context = {
        'activity': activity,
        'breakdown_total': breakdown_total,
        'amounts_match': amounts_match
    }
    return render(request, 'activities/procurement_detail.html', context)


@login_required
def export_procurement_excel(request):
    """Export procurement activities to Excel"""
    if not can_view_activities(request.user):
        return HttpResponseForbidden("You do not have permission to view activities")
    
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter
    
    # Filter activities with procurement
    activities = Activity.objects.filter(
        Q(is_procurement=True) | Q(has_partial_procurement=True), deleted=False
    ).select_related('status', 'currency', 'responsible_officer').prefetch_related('clusters', 'funders').order_by('-year', 'activity_id')
    
    # Create workbook
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Procurement Activities"
    
    # Define header style
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    
    # Headers
    headers = [
        'Activity ID', 'Activity Name', 'Year', 'Status', 'Clusters', 
        'Procurement Type', 'Total Budget', 'Procurement Amount', 
        'Disbursed Amount', 'Balance', 'Responsible Officer', 
        'Procurement Details'
    ]
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Data rows
    row_num = 2
    for activity in activities:
        # Get clusters as comma-separated string
        clusters_str = ', '.join([c.short_name for c in activity.clusters.all()])
        
        # Get responsible officer name
        officer = ''
        if activity.responsible_officer:
            officer = activity.responsible_officer.get_full_name() or activity.responsible_officer.username
        
        # Procurement details from breakdown
        procurement_details = ''
        if activity.procurement_breakdowns:
            details_list = []
            for item in activity.procurement_breakdowns:
                item_desc = f"{item.get('description', 'N/A')} - {activity.currency.code if activity.currency else 'ZMK'} {item.get('amount', 0):,.2f}"
                details_list.append(item_desc)
            procurement_details = '; '.join(details_list)
        
        # Write row
        ws.cell(row=row_num, column=1).value = activity.activity_id
        ws.cell(row=row_num, column=2).value = activity.name
        ws.cell(row=row_num, column=3).value = activity.year
        ws.cell(row=row_num, column=4).value = activity.status.name
        ws.cell(row=row_num, column=5).value = clusters_str
        ws.cell(row=row_num, column=6).value = activity.get_procurement_type_display() if activity.procurement_type else 'N/A'
        ws.cell(row=row_num, column=7).value = float(activity.total_budget or 0)
        ws.cell(row=row_num, column=8).value = float(activity.procurement_amount or 0)
        ws.cell(row=row_num, column=9).value = float(activity.disbursed_amount or 0)
        ws.cell(row=row_num, column=10).value = float(activity.balance() or 0)
        ws.cell(row=row_num, column=11).value = officer
        ws.cell(row=row_num, column=12).value = procurement_details
        
        row_num += 1
    
    # Adjust column widths
    column_widths = [15, 50, 10, 20, 25, 20, 15, 18, 18, 15, 25, 60]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    # Set number format for currency columns
    for row in range(2, row_num):
        for col in [7, 8, 9, 10]:  # Budget columns
            ws.cell(row=row, column=col).number_format = '#,##0.00'
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=procurement_activities_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    
    wb.save(response)
    return response


@login_required
def export_activities_excel(request):
    """Export filtered activities list to Excel with comprehensive details."""
    if not can_view_activities(request.user):
        return HttpResponseForbidden("You do not have permission to export activities")

    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    from openpyxl.utils import get_column_letter

    qs = (
        Activity.objects.filter(deleted=False)
        .select_related('status', 'currency', 'responsible_officer', 'procurement_type', 'parent_activity')
        .prefetch_related('clusters', 'funders')
        .order_by('-year', 'activity_id')
    )

    available_years = sorted(
        Activity.objects.filter(deleted=False).values_list('year', flat=True).distinct(),
        reverse=True
    )

    year = request.GET.get('year')
    cluster = request.GET.get('cluster')
    funder = request.GET.get('funder')
    status = request.GET.get('status')
    quarter = request.GET.get('quarter')
    month = request.GET.get('month')
    search_query = request.GET.get('q')
    procurement_status = request.GET.get('procurement_status', 'all')
    recurring_filter = request.GET.get('recurring', 'all')
    show_all_years = request.GET.get('all_years') == '1'

    current_year = date.today().year
    if not year and not show_all_years and current_year in available_years:
        year = str(current_year)
        qs = qs.filter(year=current_year)
    elif year:
        try:
            qs = qs.filter(year=int(year))
        except ValueError:
            pass

    qs = apply_role_based_filters(qs, request.user)
    qs = qs.filter(deleted=False)

    needs_distinct = False

    if cluster:
        if cluster.isdigit():
            qs = qs.filter(clusters__id=int(cluster))
        else:
            qs = qs.filter(clusters__short_name=cluster)
        needs_distinct = True

    if funder:
        if funder.isdigit():
            qs = qs.filter(funders__id=int(funder))
        else:
            qs = qs.filter(funders__code=funder)
        needs_distinct = True

    if status:
        if status.isdigit():
            qs = qs.filter(status__id=int(status))
        else:
            qs = qs.filter(status__name=status)

    if quarter:
        try:
            quarter_num = int(quarter)
            if quarter_num in [1, 2, 3, 4]:
                qs = qs.filter(quarter=quarter_num)
        except ValueError:
            pass

    if month:
        try:
            qs = qs.filter(planned_month__month=int(month))
        except ValueError:
            pass

    if search_query:
        qs = qs.filter(Q(name__icontains=search_query) | Q(activity_id__icontains=search_query))

    if needs_distinct:
        qs = qs.distinct()

    if procurement_status == 'procurement_only':
        qs = qs.filter(is_procurement=True)
    elif procurement_status == 'has_procurement':
        qs = qs.filter(Q(is_procurement=True) | Q(has_partial_procurement=True))
    elif procurement_status == 'non_procurement':
        qs = qs.filter(is_procurement=False, has_partial_procurement=False)

    if recurring_filter == 'recurring_only':
        qs = qs.filter(is_recurring=True)
    elif recurring_filter == 'generated':
        qs = qs.filter(generated_from_recurrence=True)
    elif recurring_filter == 'non_recurring':
        qs = qs.filter(is_recurring=False, generated_from_recurrence=False)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Activities"

    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    headers = [
        'Activity ID', 'Activity Name', 'Category', 'Year', 'Status', 'Planned Month', 'Quarter',
        'Clusters', 'Funders', 'Currency', 'Total Budget', 'Disbursed Amount', 'Balance',
        'Responsible Officer', 'Fully Implemented By', 'Actual Start Date', 'Actual Completion Date',
        'Retired', 'Technical Report Available', 'Notes',
        'Is Recurring', 'Recurrence Pattern', 'Recurrence Interval', 'Recurrence End Date',
        'Generated From Recurrence', 'Parent Activity ID',
        'Is Procurement', 'Has Partial Procurement', 'Procurement Type', 'Procurement Amount', 'Procurement Breakdown'
    ]

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment

    row_num = 2
    for activity in qs:
        category_str = ', '.join(activity.get_category_display_list()) if activity.category else ''
        clusters_str = ', '.join([c.short_name for c in activity.clusters.all()])
        funders_str = ', '.join([f.name for f in activity.funders.all()])
        officer = ''
        if activity.responsible_officer:
            officer = activity.responsible_officer.get_full_name() or activity.responsible_officer.username

        procurement_breakdown = ''
        if activity.procurement_breakdowns and isinstance(activity.procurement_breakdowns, list):
            breakdown_lines = []
            for item in activity.procurement_breakdowns:
                if isinstance(item, dict):
                    item_type = item.get('type', '')
                    amount = item.get('amount', 0)
                    description = item.get('description', '')
                    breakdown_lines.append(f"{item_type}: {amount} ({description})")
            procurement_breakdown = '; '.join(breakdown_lines)

        ws.cell(row=row_num, column=1).value = activity.activity_id
        ws.cell(row=row_num, column=2).value = activity.name
        ws.cell(row=row_num, column=3).value = category_str
        ws.cell(row=row_num, column=4).value = activity.year
        ws.cell(row=row_num, column=5).value = activity.status.name if activity.status else ''
        ws.cell(row=row_num, column=6).value = activity.planned_month.strftime('%Y-%m-%d') if activity.planned_month else ''
        ws.cell(row=row_num, column=7).value = activity.quarter or ''
        ws.cell(row=row_num, column=8).value = clusters_str
        ws.cell(row=row_num, column=9).value = funders_str
        ws.cell(row=row_num, column=10).value = activity.currency.code if activity.currency else 'ZMK'
        ws.cell(row=row_num, column=11).value = float(activity.total_budget or 0)
        ws.cell(row=row_num, column=12).value = float(activity.disbursed_amount or 0)
        ws.cell(row=row_num, column=13).value = float(activity.balance() or 0)
        ws.cell(row=row_num, column=14).value = officer
        ws.cell(row=row_num, column=15).value = activity.fully_implemented_by.strftime('%Y-%m-%d') if activity.fully_implemented_by else ''
        ws.cell(row=row_num, column=16).value = activity.actual_start_date.strftime('%Y-%m-%d') if activity.actual_start_date else ''
        ws.cell(row=row_num, column=17).value = activity.actual_completion_date.strftime('%Y-%m-%d') if activity.actual_completion_date else ''
        ws.cell(row=row_num, column=18).value = 'Yes' if activity.retired else 'No'
        ws.cell(row=row_num, column=19).value = 'Yes' if activity.technical_report_available else 'No'
        ws.cell(row=row_num, column=20).value = activity.notes or ''
        ws.cell(row=row_num, column=21).value = 'Yes' if activity.is_recurring else 'No'
        ws.cell(row=row_num, column=22).value = activity.get_recurrence_pattern_display() if activity.recurrence_pattern else ''
        ws.cell(row=row_num, column=23).value = activity.recurrence_interval or ''
        ws.cell(row=row_num, column=24).value = activity.recurrence_end_date.strftime('%Y-%m-%d') if activity.recurrence_end_date else ''
        ws.cell(row=row_num, column=25).value = 'Yes' if activity.generated_from_recurrence else 'No'
        ws.cell(row=row_num, column=26).value = activity.parent_activity.activity_id if activity.parent_activity else ''
        ws.cell(row=row_num, column=27).value = 'Yes' if activity.is_procurement else 'No'
        ws.cell(row=row_num, column=28).value = 'Yes' if activity.has_partial_procurement else 'No'
        ws.cell(row=row_num, column=29).value = activity.procurement_type.name if activity.procurement_type else ''
        ws.cell(row=row_num, column=30).value = float(activity.procurement_amount or 0)
        ws.cell(row=row_num, column=31).value = procurement_breakdown

        row_num += 1

    column_widths = [
        15, 45, 22, 8, 20, 14, 8, 24, 28, 10, 16, 16, 14,
        24, 16, 16, 18, 10, 16, 45,
        12, 18, 16, 16, 18, 16,
        14, 20, 18, 16, 55
    ]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for row in range(2, row_num):
        for col in [11, 12, 13, 30]:
            ws.cell(row=row, column=col).number_format = '#,##0.00'

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{max(row_num - 1, 1)}"

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = (
        f'attachment; filename=activities_export_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

    wb.save(response)
    return response

